import streamlit as st
import json
import math
import pandas as pd
import io
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# --- App Configuration ---
st.set_page_config(page_title="Mould Coordinate Generator", page_icon="🍫", layout="centered")

# --- Globals & Session State ---
COLS = [chr(i) for i in range(ord('A'), ord('O') + 1)] # A to O

# Streamlit re-runs the script on every interaction. 
# We use session_state to "remember" the calibration grid between uploads.
if 'CALIBRATED_GRID' not in st.session_state:
    st.session_state.CALIBRATED_GRID = {}

# --- Core Logic Functions ---
def infer_grid_coordinates(cx, cy):
    # If a calibration file hasn't been uploaded, fallback to the basic math division
    if not st.session_state.CALIBRATED_GRID:
        col_w, row_h = 1120 / 15, 620 / 8
        col_idx = int(cx // col_w)
        row_idx = int(cy // row_h) + 1
        col_idx = max(0, min(col_idx, 14))
        row_idx = max(1, min(row_idx, 8))
        return f"{COLS[col_idx]}{row_idx}"

    # Nearest neighbor search against the 120 calibrated grid points
    closest_cell = None
    min_dist = float('inf')
    for cell_id, (cal_x, cal_y) in st.session_state.CALIBRATED_GRID.items():
        dist = math.hypot(cx - cal_x, cy - cal_y)
        if dist < min_dist:
            min_dist = dist
            closest_cell = cell_id

    return closest_cell

def calibrate_from_json(json_text):
    try:
        if "canvas" in json_text:
            json_text = json_text.replace(' canvas"', '"').replace('canvas', '')
        data = json.loads(json_text)
        image_metadata = data.get("_via_img_metadata", data)
    except Exception as e:
        st.error(f"❌ Error: Failed to parse calibration payload. Details: {e}")
        return False

    all_cavities = []
    first_image_key = list(image_metadata.keys())[0]
    regions = image_metadata[first_image_key].get("regions", [])

    for item in regions:
        shape = item["shape_attributes"]
        if "x" in shape:
            cx = int(round(shape["x"] + (shape["width"] / 2)))
            cy = int(round(shape["y"] + (shape["height"] / 2)))
            all_cavities.append((cx, cy))

    if len(all_cavities) != 120:
        st.warning(f"⚠ Warning: Calibration requires exactly 120 boxes. Found {len(all_cavities)} in the file.")
        return False

    all_cavities.sort(key=lambda c: c[0])
    st.session_state.CALIBRATED_GRID.clear()

    for col_idx in range(15):
        col_letter = COLS[col_idx]
        col_chunk = all_cavities[col_idx * 8 : (col_idx + 1) * 8]
        col_chunk.sort(key=lambda c: c[1])

        for row_idx, (cx, cy) in enumerate(col_chunk):
            cell_id = f"{col_letter}{row_idx + 1}"
            st.session_state.CALIBRATED_GRID[cell_id] = (cx, cy)

    st.success("✅ Grid successfully calibrated with 120 exact reference points (A1 to O8)!")
    return True

# --- Excel Styling Functions ---
def apply_header_styling(worksheet, max_column):
    navy_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    white_bold_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(bottom=Side(style='medium', color="000000"))

    for col in range(1, max_column + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = navy_fill
        cell.font = white_bold_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    worksheet.row_dimensions[1].height = 24

def autofit_columns(worksheet):
    for col in worksheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1 and worksheet.title == 'Coordinate Counts Ranking':
                continue
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

def save_with_merged_cells(writer, df, sheet_name):
    workbook = writer.book
    worksheet = workbook.create_sheet(title=sheet_name)

    for r in dataframe_to_rows(df, index=False, header=True):
        worksheet.append(r)

    apply_header_styling(worksheet, df.shape[1])
    thin_gray = Side(style='thin', color='D9D9D9')
    cell_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for row in range(2, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            worksheet.cell(row=row, column=col).border = cell_border
            if col > 1:
                worksheet.cell(row=row, column=col).alignment = Alignment(horizontal="center")

    start_row = 2
    max_row = worksheet.max_row
    while start_row <= max_row:
        current_val = worksheet.cell(row=start_row, column=1).value
        end_row = start_row
        while end_row + 1 <= max_row and worksheet.cell(row=end_row + 1, column=1).value == current_val:
            end_row += 1

        if end_row > start_row:
            worksheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            merged_cell = worksheet.cell(row=start_row, column=1)
            merged_cell.alignment = Alignment(vertical="center", horizontal="left")
        start_row = end_row + 1

    autofit_columns(worksheet)

def create_visual_map(workbook, summary_df):
    map_ws = workbook.create_sheet(title='Visual Mould Map')

    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    empty_mould_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    empty_mould_font = Font(color="9C0006", bold=True)
    good_mould_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    good_mould_font = Font(color="006100")

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin', color='D9D9D9'),
                         right=Side(style='thin', color='D9D9D9'),
                         top=Side(style='thin', color='D9D9D9'),
                         bottom=Side(style='thin', color='D9D9D9'))

    for i, col_letter in enumerate(COLS):
        cell = map_ws.cell(row=1, column=i+2, value=col_letter)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        map_ws.column_dimensions[get_column_letter(i+2)].width = 12

    for r in range(1, 9):
        cell = map_ws.cell(row=r+1, column=1, value=r)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        map_ws.row_dimensions[r+1].height = 40
        map_ws.column_dimensions['A'].width = 6

    summary_dict = dict(zip(summary_df['Grid_Coordinate'], summary_df['Total_Count']))

    for r in range(1, 9):
        for c_idx, col_letter in enumerate(COLS):
            coord = f"{col_letter}{r}"
            cell = map_ws.cell(row=r+1, column=c_idx+2)
            cell.alignment = center_align
            cell.border = thin_border

            if coord in summary_dict:
                count = summary_dict[coord]
                cell.value = f"X\n({count})"
                cell.fill = empty_mould_fill
                cell.font = empty_mould_font
            else:
                cell.value = "OK"
                cell.fill = good_mould_fill
                cell.font = good_mould_font

def process_raw_json(json_text):
    try:
        if "canvas" in json_text:
            json_text = json_text.replace(' canvas"', '"').replace('canvas', '')
        data = json.loads(json_text)
        image_metadata = data.get("_via_img_metadata", data)
    except Exception as e:
        st.error(f"❌ Error: Failed to parse file payload. Details: {e}")
        return None

    all_output_rows = []
    total_moulds = 0

    for unique_id, mould_profile in image_metadata.items():
        filename = mould_profile.get("filename", unique_id)
        regions = mould_profile.get("regions", [])

        total_moulds += 1
        if not regions:
            continue

        mould_cavities = []
        for idx, item in enumerate(regions):
            shape = item["shape_attributes"]
            if "x" in shape:
                cx = int(round(shape["x"] + (shape["width"] / 2)))
                cy = int(round(shape["y"] + (shape["height"] / 2)))
                grid_cell = infer_grid_coordinates(cx, cy)
                mould_cavities.append({"cx": cx, "cy": cy, "inferred": grid_cell})

        mould_cavities = sorted(mould_cavities, key=lambda val: (val["cy"], val["cx"]))

        for relative_idx, cavity in enumerate(mould_cavities):
            all_output_rows.append({
                "Mould_File": filename,
                "Box_Index": relative_idx + 1,
                "Center_X": cavity["cx"],
                "Center_Y": cavity["cy"],
                "Grid_Cell": cavity["inferred"]
            })

    if all_output_rows:
        final_df = pd.DataFrame(all_output_rows)
        summary_df = final_df['Grid_Cell'].value_counts().reset_index()
        summary_df.columns = ['Grid_Coordinate', 'Total_Count']
        summary_df['Occurrence_Rate'] = summary_df['Total_Count'] / total_moulds

        # Write to memory instead of a file on disk
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            save_with_merged_cells(writer, final_df, 'Detailed Coordinates')

            workbook = writer.book
            summary_ws = workbook.create_sheet(title='Coordinate Counts Ranking')

            summary_ws.merge_cells('A1:C1')
            banner_cell = summary_ws['A1']
            banner_cell.value = f"📊 Total Unique Molds Evaluated: {total_moulds}"
            banner_cell.font = Font(name="Calibri", size=11, bold=True, color="1F497D")
            banner_cell.fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
            banner_cell.alignment = Alignment(horizontal="left", vertical="center")
            summary_ws.row_dimensions[1].height = 28

            for col_idx, col_name in enumerate(summary_df.columns, start=1):
                summary_ws.cell(row=2, column=col_idx, value=col_name)

            navy_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            white_bold_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            for col in range(1, summary_df.shape[1] + 1):
                c = summary_ws.cell(row=2, column=col)
                c.fill = navy_fill
                c.font = white_bold_font
                c.alignment = Alignment(horizontal="center", vertical="center")
            summary_ws.row_dimensions[2].height = 24

            for r_idx, row_data in enumerate(dataframe_to_rows(summary_df, index=False, header=False), start=3):
                for c_idx, val in enumerate(row_data, start=1):
                    summary_ws.cell(row=r_idx, column=c_idx, value=val)

            thin_gray = Side(style='thin', color='D9D9D9')
            cell_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

            for row in range(3, summary_ws.max_row + 1):
                for col in range(1, summary_ws.max_column + 1):
                    cell = summary_ws.cell(row=row, column=col)
                    cell.border = cell_border
                    cell.alignment = Alignment(horizontal="center")
                    if col == 3:
                        cell.number_format = '0.0%'

            autofit_columns(summary_ws)
            create_visual_map(workbook, summary_df)

        return output.getvalue()
    else:
        st.warning("⚠ Processed, but no valid bounding box annotations were found in this file.")
        return None

# --- Streamlit User Interface ---

st.title("🍫 Chocolate Mould Coordinate Generator")
st.markdown("---")

st.header("Step 1: Calibration (Recommended)")
st.markdown("Upload your 120-box master JSON to calibrate exact positioning. If skipped, standard math grids apply.")
cal_file = st.file_uploader("Upload Calibration JSON", type=["json"], key="cal")

if cal_file is not None:
    file_content = cal_file.getvalue().decode("utf-8")
    calibrate_from_json(file_content)

st.markdown("---")

st.header("Step 2: Generate Analysis")
st.markdown("Upload your exported VIA JSON files to process the empty molds.")

if st.session_state.CALIBRATED_GRID:
    st.info("✨ Using calibrated 120-point coordinate grid.")
else:
    st.info("⚠ Using default mathematical grid.")

analysis_file = st.file_uploader("Upload Analysis JSON", type=["json"], key="analysis")

if analysis_file is not None:
    with st.spinner("Processing file and building spreadsheet..."):
        file_content = analysis_file.getvalue().decode("utf-8")
        excel_data = process_raw_json(file_content)
        
        if excel_data:
            st.success("🎉 Processing successful! Your spreadsheet is ready.")
            st.download_button(
                label="📥 Download Excel Report",
                data=excel_data,
                file_name="mould_cavity_analysis.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
